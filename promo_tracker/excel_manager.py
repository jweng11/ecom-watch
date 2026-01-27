"""
Excel management for promotion tracking.
Maintains a structured, searchable Excel file with promotion history.
"""
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


# Excel column definitions
COLUMNS = [
    ("scrape_date", "Scrape Date", 12),
    ("retailer", "Retailer", 12),
    ("product_name", "Product Name", 40),
    ("brand", "Brand", 12),
    ("original_price", "Original Price", 14),
    ("sale_price", "Sale Price", 12),
    ("discount_amount", "Discount $", 12),
    ("discount_percent", "Discount %", 12),
    ("promo_type", "Promo Type", 15),
    ("key_specs", "Key Specs", 45),
    ("promo_end_date", "Promo Ends", 12),
    ("screenshot_path", "Screenshot", 50),
    ("source_url", "Source URL", 50),
]


def create_workbook(filepath: Path) -> openpyxl.Workbook:
    """Create a new workbook with formatted headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Promotions"

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Write headers
    for col_idx, (key, display_name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=display_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add filter
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(COLUMNS))}1"

    wb.save(filepath)
    return wb


def load_or_create_workbook(filepath: Path) -> openpyxl.Workbook:
    """Load existing workbook or create new one."""
    if filepath.exists():
        return openpyxl.load_workbook(filepath)
    else:
        filepath.parent.mkdir(parents=True, exist_ok=True)
        return create_workbook(filepath)


def add_promotions(
    filepath: Path,
    promotions: list[dict],
    scrape_date: str,
    retailer: str,
    screenshot_path: str,
    source_url: str
) -> int:
    """
    Add promotions to the Excel file.

    Args:
        filepath: Path to Excel file
        promotions: List of promotion dicts from LLM
        scrape_date: Date of scrape (YYYY-MM-DD)
        retailer: Retailer name
        screenshot_path: Path to screenshot file
        source_url: URL that was scraped

    Returns:
        Number of promotions added
    """
    wb = load_or_create_workbook(filepath)
    ws = wb.active

    # Find next empty row
    next_row = ws.max_row + 1

    # Style for data cells
    data_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Alternating row colors
    light_fill = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    added_count = 0

    for promo in promotions:
        row_fill = light_fill if next_row % 2 == 0 else white_fill

        row_data = [
            scrape_date,
            retailer,
            promo.get("product_name", ""),
            promo.get("brand", ""),
            promo.get("original_price", ""),
            promo.get("sale_price", ""),
            promo.get("discount_amount", ""),
            promo.get("discount_percent", ""),
            promo.get("promo_type", ""),
            promo.get("key_specs", ""),
            promo.get("promo_end_date", ""),
            screenshot_path,
            source_url,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.fill = row_fill

        next_row += 1
        added_count += 1

    # Update filter range
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(COLUMNS))}{next_row - 1}"

    wb.save(filepath)
    return added_count


def create_summary_sheet(filepath: Path):
    """Add or update a summary sheet with aggregated stats."""
    wb = openpyxl.load_workbook(filepath)

    # Remove existing summary sheet if present
    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    ws_summary = wb.create_sheet("Summary", 0)
    ws_data = wb["Promotions"]

    # Gather stats
    total_rows = ws_data.max_row - 1  # Exclude header

    if total_rows == 0:
        ws_summary.cell(row=1, column=1, value="No data yet")
        wb.save(filepath)
        return

    # Count by retailer
    retailer_counts = {}
    brand_counts = {}
    dates = set()

    for row in range(2, ws_data.max_row + 1):
        retailer = ws_data.cell(row=row, column=2).value
        brand = ws_data.cell(row=row, column=4).value
        date = ws_data.cell(row=row, column=1).value

        retailer_counts[retailer] = retailer_counts.get(retailer, 0) + 1
        if brand:
            brand_counts[brand] = brand_counts.get(brand, 0) + 1
        if date:
            dates.add(str(date))

    # Style
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)

    # Write summary
    row = 1
    ws_summary.cell(row=row, column=1, value="Laptop Promotion Tracker Summary").font = header_font
    row += 2

    ws_summary.cell(row=row, column=1, value=f"Total Promotions Tracked: {total_rows}")
    row += 1
    ws_summary.cell(row=row, column=1, value=f"Scrape Dates: {len(dates)}")
    row += 2

    ws_summary.cell(row=row, column=1, value="By Retailer:").font = subheader_font
    row += 1
    for retailer, count in sorted(retailer_counts.items(), key=lambda x: -x[1]):
        ws_summary.cell(row=row, column=1, value=f"  {retailer}: {count}")
        row += 1
    row += 1

    ws_summary.cell(row=row, column=1, value="By Brand:").font = subheader_font
    row += 1
    for brand, count in sorted(brand_counts.items(), key=lambda x: -x[1])[:10]:
        ws_summary.cell(row=row, column=1, value=f"  {brand}: {count}")
        row += 1

    ws_summary.column_dimensions["A"].width = 50

    wb.save(filepath)
